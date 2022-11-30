import os
import shutil
from converter import *
import PySimpleGUI as sg
import openpyxl
import csv
from itertools import zip_longest

sg.theme('LightBrown1')
font = ("Arial", 16)
button_color = "grey"

folderToTracklist = os.path.join(os.path.join(os.path.expanduser('~')), 'Downloads')
desktop = os.path.join(os.path.join(os.path.expanduser('~')), 'Desktop')

link = []
song = []
klasse = []

# gui layout
layout1 = [[sg.T("        ", size=(15,3))],
        [sg.Text('Pfad MIXFERTIG:', font=font)],
        [sg.Input(), sg.FolderBrowse('Browse', key='-dirNAS-', initial_folder=desktop, size=(12,1))],
        [sg.Text('Schul-ID:', font=font)],
        [sg.InputText(key='-schoolID-')],
        [sg.Text('Schul-Name:', font=font)],
        [sg.InputText(key='-schoolName-')],
        [sg.Text('Pfad TRACKLIST ADMINTOOL:', font=font)], 
        [sg.Input(), sg.FileBrowse('Browse', key='-dirTL-', initial_folder=folderToTracklist, size=(12,1))],
        [sg.T("        ")],
        [sg.Radio('Ganzjahr', "RADIO1", default=False, key="-RADIO1-", font=font), sg.Radio('Xmas', "RADIO1", default=True, key="-RADIO2-", font=font)],
        [sg.T("        ", size=(10,14))],
        [sg.Button('OK!', key="-START-", size=(15,1.2)), sg.Button('Cancel', key="-CANCEL-", size=(15,1.2))]]

layout2 = [[sg.T("        ", size=(15,3))],
        [sg.Text('Schul-ID:', font=font)],
        [sg.InputText(key='-schoolID2-')],
        [sg.Text('Pfad TRACKLIST ADMINTOOL:', font=font)], 
        [sg.Input(), sg.FileBrowse('Browse', key='-dirTL2-', initial_folder=folderToTracklist, size=(12,1))],
        [sg.T("        ", size=(10,28))],
        [sg.Button('OK!', key="-START2-", size=(15,1.2)), sg.Button('Cancel', key="-CANCEL2-", size=(15,1.2))]]

layout3 = [[sg.T("        ")],
        [sg.Checkbox('Mix fertig?', default=False, font=font)],
        [sg.Checkbox('Download tracklist.xlsx aus Admintool', default=False, font=font)],
        [sg.Checkbox('WAV-Check -> DDP erstellen?', default=False, font=font)],
        [sg.Checkbox('tracklist.txt in Filemaker kopieren', default=False, font=font)],
        [sg.Checkbox('Dropbox-Check', default=False, font=font)],
        [sg.Checkbox('Transmit-Check', default=False, font=font)],
        [sg.T("  ---------------------------------------------------------------------", font=font)],
        [sg.Text('Dateien aus Filemaker in Dropbox', font=font)],
        [sg.Checkbox('codes.csv', default=False, font=font)], 
        [sg.Checkbox('minicards.pdf', default=False, font=font)],
        [sg.Checkbox('booklet.pdf', default=False, font=font)],
        [sg.Checkbox('label.pdf', default=False, font=font)],
        [sg.T("  ---------------------------------------------------------------------", font=font)],
        [sg.Text('Uploads Admintool', font=font)],
        [sg.Checkbox('codes.csv', default=False, font=font)], 
        [sg.Checkbox('tracklist.csv', default=False, font=font)],
        [sg.T("        ")]
        ]

tabgrp = [[sg.TabGroup([[sg.Tab('       dave        ', layout1), sg.Tab('  tracklist only  ', layout2), sg.Tab('    checklist   ', layout3)]], font=font,
                       title_color='grey', tab_background_color='lightgrey',selected_title_color='white',
                       selected_background_color='grey', border_width=5)]] 

window = sg.Window('DAVE', tabgrp)

# "/Users/horthin/Documents/scripts/LM_DAVE/dave_logo.txt"
with open(("./dave_logo.txt"), 'r') as f:
    file_content = f.read()
    print(file_content,"\nWaiting for input...")

while True:
    event, values = window.read()

    dirNAS = values['-dirNAS-']
    schoolID = values['-schoolID-']
    schoolName = values['-schoolName-']
    tracklist = values['-dirTL-']
    schoolID2 = values['-schoolID2-']
    tracklist2 = values['-dirTL2-']

    if event == sg.WIN_CLOSED or event == '-CANCEL-' or event == '-CANCEL2-': # if user closes window or clicks cancel
        print("abort. see you next time!")
        break

    if event == "-START2-":
        if schoolID2 == "":
            print('Keine School-ID')
        else:
            try:
                schoolID2 = int(schoolID2)
                schoolID2 = str(schoolID2).zfill(4)
                print(schoolID2)
            except OSError:
                pass

        if tracklist2 == "":
            print('Kein Pfad -> Tracklist!')
        else:
            # # Excel Tabelle Admin-Tool lesen
            excel_file = openpyxl.load_workbook(tracklist2)
            sheet_obj = excel_file.active
            m_row = sheet_obj.max_row

            # Tracklist-Daten
            txtFile = schoolID2 + "_tracklist.txt"
            txt_file = open(txtFile, "w")

            j = 1

            for i in range(2, m_row + 1):
                cell_obj = sheet_obj.cell(row = i, column = 1)
                song.append(cell_obj.value)
                cell_obj2 = sheet_obj.cell(row = i, column = 2)
                klasse.append(cell_obj2.value)
                try:
                    full_row = str(j).zfill(2) + " " + cell_obj.value + " " + "(" + cell_obj2.value + ")\n" # optional for copy from txt-file, saved in original path
                    txt_file.write(str(full_row)) # write to .txt if 3 columns (titel + (klasse))
                except:
                    full_row2 = str(j).zfill(2) + " " + cell_obj.value + "\n"
                    txt_file.write(str(full_row2)) # write to .txt if 2 columns (alle kinder singen alle songs)
                
                j += 1

            txt_file.close()
            shutil.move(("./" + txtFile), desktop)
            print("check desktop for tracklist!")

            break
        continue

    dirNAS = values['-dirNAS-']
    if dirNAS == "":
        print('Kein Pfad -> MIXFERTIG!')

    if schoolID == "":
        print('Keine School-ID')
    else:
        try:
            schoolIDint = int(values['-schoolID-'])
            schoolID = str(schoolID).zfill(4)
            cache_folder = desktop + "/cache_MM" + schoolID
            print(schoolID)
            print(cache_folder)
        except OSError:
            pass
        
    schoolName = values['-schoolName-']
    if schoolName == "":
        print("Kein Schulname!")

    tracklist = values['-dirTL-']
    if tracklist == "":
        print('Kein Pfad -> Tracklist!')

    if event == '-START-' and dirNAS != "" and schoolID != "" and tracklist != "":

        # # Excel Tabelle Admin-Tool lesen
        excel_file = openpyxl.load_workbook(tracklist)
        sheet_obj = excel_file.active
        m_row = sheet_obj.max_row

        # Tracklist-Daten
        txtFile = schoolID + "_tracklist.txt"
        txt_file = open(txtFile, "w")

        j = 1

        for i in range(2, m_row + 1):
            cell_obj = sheet_obj.cell(row = i, column = 1)
            song.append(cell_obj.value)
            cell_obj2 = sheet_obj.cell(row = i, column = 2)
            klasse.append(cell_obj2.value)
            try:
                full_row = str(j).zfill(2) + " " + cell_obj.value + " " + "(" + cell_obj2.value + ")\n" # optional for copy from txt-file, saved in original path
                txt_file.write(str(full_row)) # write to .txt if 3 columns (titel + (klasse))
            except:
                full_row2 = str(j).zfill(2) + " " + cell_obj.value + "\n"
                txt_file.write(str(full_row2)) # write to .txt if 2 columns (alle kinder singen alle songs)
            
            j += 1

        txt_file.close()
        shutil.move(("./" + txtFile), desktop)

        # # create cache folder on desktop with cache_schoolID
        try:
            if not os.path.exists(cache_folder):
                os.mkdir(cache_folder)
        except OSError:
            print(cache_folder, "konnte nicht erstellt werden!")

        # # copy wav files to cache folder on desktop
        try:
            for file in os.listdir(dirNAS):
                if file.endswith(".wav") == True:
                    wav_file = dirNAS + "/" + file
                    shutil.copy(wav_file, cache_folder)       
        except:
            print("kein pfad vorhanden")

        print("\n-------------------------------------------------------------")
        print("1. wav check!")
        print("2. copy tracklist to clipboard!")
        print("\n... waiting for user input ...")
        userInput = input("3. go? [yes|no] ")
        print("-------------------------------------------------------------\n")

        if userInput == "no" or userInput == "No" or userInput == "NO" or userInput == "n":
            # # delete cache Folder 
            try:
                shutil.rmtree(cache_folder)
                os.remove(desktop + "/" + txtFile)
            except OSError as e:
                print("Error: %s - %s." % (e.filename, e.strerror))
            print("cache_folder and txt-file on desktop deleted. see you next time!")
            break

        elif userInput == "yes" or userInput == "Yes" or userInput == "YES" or userInput == "y":
            # # Convert wav to mp3
            conv = mp3_converter(cache_folder, ".wav", "mp3")
            conv.lower_underscore()
            conv.mp3()

            # # Rename mp3s in cache_schoolID and move to /mp3
            mp3_folder = cache_folder + "/mp3"
            if not os.path.exists(mp3_folder):
                os.makedirs(mp3_folder)
            for filename in os.listdir(cache_folder):
                if (filename.endswith(".mp3")):
                    source = os.path.join(cache_folder, filename)
                    shutil.move(source, mp3_folder)
            os.listdir(mp3_folder)
            for f in os.listdir(mp3_folder):
                os.rename(os.path.join(mp3_folder, f), os.path.join(mp3_folder, f).replace('§', ' ').title().replace('.Mp3', '.mp3'))

            # # Rename wavs in cache_schoolID and move to /wav
            wav_folder = cache_folder + "/wav"
            if not os.path.exists(wav_folder):
                os.makedirs(wav_folder)
            for filename in os.listdir(cache_folder):
                if (filename.endswith(".wav")):
                    source = os.path.join(cache_folder, filename)
                    shutil.move(source, wav_folder)
            os.listdir(wav_folder)
            for f in os.listdir(wav_folder):
                os.rename(os.path.join(wav_folder, f), os.path.join(wav_folder, f).replace('§', ' ').title().replace('.Wav', '.wav'))

            # # zip mp3 files
            os.path.join(mp3_folder, shutil.make_archive("AlleLieder", 'zip', mp3_folder))
            shutil.move(("./AlleLieder.zip"), mp3_folder)

            # # build csv and txt with links to upload to dropbox
            while True:
                lines = os.listdir(mp3_folder)
                for mp3s in lines:
                    link.append("https://www.hoerthin.de/mp3/" + schoolID + "/" + mp3s.replace(" ", "%20"))
                link.sort()

                csvFile = schoolID + "_mp3Liste.csv"

                song.append("Alle Lieder")
                klasse.append(" ")
                if event == "-RADIO1-":
                    link.append("https://www.hoerthin.de/mp3/Minimusikersong.mp3")
                    song.append("Minimusikersong")
                else:
                    link.append("https://www.hoerthin.de/mp3/Minimusikersong%20Xmas.mp3")
                    song.append("Minimusikersong Xmas")
                klasse.append("Minimusiker")

                data = [link, song, klasse]
                export_data = zip_longest(*data, fillvalue = '')
                with open(csvFile, 'w', newline='') as file:
                    write = csv.writer(file)
                    write.writerows(export_data)

                break
            # txt_file.close()
            file.close()

            # upload files to mysql -> file using customer_id(schoolID), event_id
            # soon to come

            # # copy csv and txt files to cache folder
            shutil.move((desktop + "/" + txtFile), cache_folder)
            shutil.move(("./" + csvFile), cache_folder)

            # # rename cache_folder+schoolID to schoolID
            for foldername in os.listdir(desktop):
                if foldername.startswith("cache_"):
                    os.rename(os.path.join(desktop, foldername), os.path.join(desktop, foldername).replace(("cache_MM"+schoolID), ("MM"+schoolID+" "+schoolName)))

            # # duplicate full folder to dropbox
            # "Macintosh HD:Users:horthin:Dropbox:Apps:FileTrip_deinecd:Uploads:__Minimusiker - xxx:"
            command1 = """ osascript -e '
            set dropboxFolder to "Macintosh HD:Users:minimusiker:Dropbox:Apps:FileTrip_deinecd:Uploads:__Minimusiker - xxx:"
            set desktopFolder to (path to desktop)
            set theFolder to "MM"
            tell application "Finder"
                activate
                set matchingFolder to ((every folder in desktop) whose name begins with theFolder) as text
                duplicate (folder matchingFolder) to dropboxFolder
            end tell
            '"""
            
            os.system(command1)
            from ftp_upload import *
            os.system("ftp_upload.py 1")

window.close()